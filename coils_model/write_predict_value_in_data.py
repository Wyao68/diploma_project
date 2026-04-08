import os
import json
import torch
import numpy as np
import pandas as pd
import openpyxl

import FC_model

base = os.path.dirname(os.path.dirname(__file__))
data_path = os.path.join(base, 'saved_models')


def load_model(state_path: str, net_dims: list[int], dropout_p: float, device: torch.device) -> torch.nn.Module:
    """
    加载模型 
    """
    model = FC_model.FullyConnectedNet(net_dims, dropout_p=dropout_p)

    # 加载模型权重
    model.load_state_dict(torch.load(state_path, map_location=device))

    return model.to(device)


def main():
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    
    # 从best_hyperparams.json中加载模型参数
    with open(os.path.join(data_path, 'best_hyperparams.json'), "r") as f:
        hyperparams = json.load(f)
    
    net_dims = [7]  # 输入特征维度，需与训练时一致
    for i in range(hyperparams['n_hidden']):
        net_dims.append(hyperparams[f'n_units_layer{i}'])
    net_dims.append(2)
    
    # model = load_model(os.path.join(data_path, 'coils_model_state_dict.pt'), net_dims, dropout_p=hyperparams['dropout_p'], device=device)
    model = load_model(os.path.join(data_path, 'transfer_model_state_dict.pt'), net_dims, dropout_p=hyperparams['dropout_p'], device=device)

    # 读取模型训练时的统计量 meta
    with open(os.path.join(data_path, 'meta.json'), "r") as f: meta = json.load(f)

    # ---- 读取 data_contrast.xlsx 并用模型预测，写回预测值 ----
    if os.path.exists(target_path):
        try:
            df = pd.read_excel(target_path, header=None)   # 使用 pandas 读取含有公式单元格的excel时必须在excel中保存公式计算结果，否则公式单元格会被 pandas 读取为 NaN
            data_vol = df.to_numpy()  
        except Exception as e:
            print('Failed to read Excel file:', e)
            return
        
        x_mean = np.array(meta['x_mean'], dtype=np.float32)
        x_std = np.array(meta['x_std'], dtype=np.float32)
        # 与训练时一致的输入列索引
        input_cols = [0,1,2,3,4,7,8] 

        X_raw = data_vol[2:, input_cols] # 从第3行开始读取输入数据，并选择指定的输入列
        X_norm = (X_raw - x_mean) / x_std         
        X_t = torch.from_numpy(X_norm.astype(np.float32)).to(device)

        model.eval()
        with torch.no_grad():
            preds = model.forward(X_t).cpu().numpy()

        # 单元格的索引从1开始
        start_row = 3  
        n_rows_pred = preds.shape[0]

        available_rows = data_vol.shape[0] - (start_row - 1)
        rows_to_write = min(available_rows, n_rows_pred)

        # print(X_raw[0, :])  # 打印原始输入以供调试
        # print(X_norm[0, :])  # 打印归一化输入以供调试
        # print(preds[0, :])  # 打印预测值以供调试

        try:
            wb = openpyxl.load_workbook(target_path)
            ws = wb.active

            col_L = 28
            col_R = 29

            for i in range(rows_to_write):
                row_idx = start_row + i
                ws.cell(row=row_idx, column=col_L, value=float(preds[i, 0]))
                ws.cell(row=row_idx, column=col_R, value=float(preds[i, 1]))

            wb.save(target_path)
            print(f'Wrote predictions successfully.')
        except Exception as e:
            print('Failed to write Excel file with openpyxl:', e)
                
    else:
        print('data_contrast.xlsx not found at', target_path)

    
if __name__ == "__main__":
    # for i in range(2,8):
    #     # 需要写入预测值的 Excel 文件路径
    #     target_path = os.path.join(base, 'data_contrast', f'N{i}.xlsx')
    #     main()
        
    target_path = os.path.join(base, 'data_contrast', 'experiment_data.xlsx')
    main()
    